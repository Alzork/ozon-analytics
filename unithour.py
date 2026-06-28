"""
Hourly updater for the unit economics workbook.

Touches only these cells:
1) "Калькулятор прибыли"!L - product volume, l
2) "Калькулятор прибыли"!V - current promo price
3) "Лист1"!L - count of clusters with sellable stock for the article in K

The workbook is an .xlsx file stored in Google Drive. By default the script
uploads the changed file back to Drive; use --dry-run to preview only.
"""

from __future__ import annotations

import argparse
import os
import re
import tempfile
import time
from io import BytesIO
from pathlib import Path

import requests
from dotenv import load_dotenv
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook

from constants import clean_offer_to_product_id, clusters_id, offer_to_sku


load_dotenv()

OZON_CLIENT_ID = os.getenv("OZON_CLIENT_ID")
OZON_API_KEY = os.getenv("OZON_API_KEY")
GOOGLE_CREDS_PATH = os.getenv("GOOGLE_CREDS") or os.getenv("GOOGLE_CREDS_PATH") or "google-creds.json"

UNITHOUR_FILE_ID = os.getenv("GOOGLE_SHEET_ID_UNITHOUR") or os.getenv("GOOGLE_SHEET_ID_UNITHOUR_TEST")
UNITHOUR_WORKSHEET = os.getenv("GOOGLE_SHEET_WORKSHEET_UNITHOUR", "Калькулятор прибыли")
UNITHOUR_LOOKUP_WORKSHEET = os.getenv("GOOGLE_SHEET_LOOKUP_WORKSHEET_UNITHOUR", "Лист1")

MAIN_HEADER_ROW = int(os.getenv("UNITHOUR_MAIN_HEADER_ROW", "16"))
LOOKUP_HEADER_ROW = int(os.getenv("UNITHOUR_LOOKUP_HEADER_ROW", "1"))

MAIN_ARTICLE_COL = "B"
MAIN_VOLUME_COL = "L"
MAIN_CLUSTER_FORMULA_COL = "S"
MAIN_PRICE_COL = "V"
LOOKUP_ARTICLE_COL = "K"
LOOKUP_CLUSTER_COUNT_COL = "L"

OZON_HEADERS = {
    "Client-Id": OZON_CLIENT_ID or "",
    "Api-Key": OZON_API_KEY or "",
    "Content-Type": "application/json",
}

DRIVE_FILE_URL = "https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
DRIVE_UPLOAD_URL = "https://www.googleapis.com/upload/drive/v3/files/{file_id}?uploadType=media"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def post_with_retry(url: str, body: dict, *, timeout: int = 60, retries: int = 5) -> dict:
    for attempt in range(retries):
        try:
            resp = requests.post(url, headers=OZON_HEADERS, json=body, timeout=timeout)
            if resp.status_code == 200:
                return resp.json()
            if resp.status_code in (429, 500, 502, 503, 504):
                time.sleep(min(30, 2 ** attempt))
                continue
            resp.raise_for_status()
        except requests.RequestException:
            if attempt == retries - 1:
                raise
            time.sleep(min(30, 2 ** attempt))
    return {}


def get_google_creds():
    scopes = [
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/spreadsheets",
    ]
    return ServiceAccountCredentials.from_json_keyfile_name(GOOGLE_CREDS_PATH, scopes)


def get_drive_token() -> str:
    creds = get_google_creds()
    return creds.get_access_token().access_token


def download_drive_xlsx(file_id: str) -> bytes:
    token = get_drive_token()
    resp = requests.get(
        DRIVE_FILE_URL.format(file_id=file_id),
        headers={"Authorization": f"Bearer {token}"},
        timeout=120,
    )
    resp.raise_for_status()
    return resp.content


def upload_drive_xlsx(file_id: str, content: bytes) -> None:
    token = get_drive_token()
    resp = requests.patch(
        DRIVE_UPLOAD_URL.format(file_id=file_id),
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": XLSX_MIME,
        },
        data=content,
        timeout=180,
    )
    resp.raise_for_status()


def save_local_backup(content: bytes) -> Path:
    backup_dir = Path(tempfile.gettempdir()) / "unithour_backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / "unithour_before_last_write.xlsx"
    backup_path.write_bytes(content)
    return backup_path


def normalize_article(raw_value) -> str:
    if raw_value in (None, ""):
        return ""
    if isinstance(raw_value, (int, float)):
        value = str(int(raw_value)) if float(raw_value).is_integer() else str(raw_value)
    else:
        value = str(raw_value)

    value = value.replace("\xa0", " ").strip().upper()
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"(?i)\s*ИР\s*$", "", value).strip()
    value = re.sub(r"(?<=\d)\.0+$", "", value)
    if not value:
        return ""

    candidates = []

    def add(candidate: str) -> None:
        candidate = candidate.strip()
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    add(value)
    add(re.sub(r"(?<=\d)/\d+", "", value).strip())

    base_match = re.match(r"^\s*(\d+)", value)
    if base_match:
        add(base_match.group(1))

    for candidate in candidates:
        if candidate in offer_to_sku:
            return candidate
    return candidates[-1] if candidates else ""


def normalize_header(value) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").replace("\xa0", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def validate_layout(workbook) -> tuple[object, object]:
    if UNITHOUR_WORKSHEET not in workbook.sheetnames:
        raise RuntimeError(f"Лист {UNITHOUR_WORKSHEET!r} не найден")
    if UNITHOUR_LOOKUP_WORKSHEET not in workbook.sheetnames:
        raise RuntimeError(f"Лист {UNITHOUR_LOOKUP_WORKSHEET!r} не найден")

    main_ws = workbook[UNITHOUR_WORKSHEET]
    lookup_ws = workbook[UNITHOUR_LOOKUP_WORKSHEET]

    expected_headers = {
        f"{MAIN_ARTICLE_COL}{MAIN_HEADER_ROW}": "артикул",
        f"{MAIN_VOLUME_COL}{MAIN_HEADER_ROW}": "объем товара, л",
        f"{MAIN_CLUSTER_FORMULA_COL}{MAIN_HEADER_ROW}": "количество кластеров, где есть товар",
        f"{MAIN_PRICE_COL}{MAIN_HEADER_ROW}": "текущая цена, руб",
    }
    for cell, expected in expected_headers.items():
        actual = normalize_header(main_ws[cell].value)
        if actual != expected:
            raise RuntimeError(f"Неожиданный заголовок {UNITHOUR_WORKSHEET}!{cell}: {actual!r}, ожидали {expected!r}")

    lookup_expected = {
        f"{LOOKUP_ARTICLE_COL}{LOOKUP_HEADER_ROW}": "артикул",
        f"{LOOKUP_CLUSTER_COUNT_COL}{LOOKUP_HEADER_ROW}": "кол-во складов",
    }
    for cell, expected in lookup_expected.items():
        actual = normalize_header(lookup_ws[cell].value)
        if actual != expected:
            raise RuntimeError(
                f"Неожиданный заголовок {UNITHOUR_LOOKUP_WORKSHEET}!{cell}: {actual!r}, ожидали {expected!r}"
            )

    return main_ws, lookup_ws


def collect_main_rows(main_ws) -> list[tuple[int, str, str]]:
    rows = []
    for row_idx in range(MAIN_HEADER_ROW + 1, main_ws.max_row + 1):
        raw_article = main_ws[f"{MAIN_ARTICLE_COL}{row_idx}"].value
        offer = normalize_article(raw_article)
        if not raw_article or not offer:
            continue
        if offer not in offer_to_sku:
            print(f"WARN main row {row_idx}: article {raw_article!r} normalized to {offer!r}, not found in offer_to_sku")
            continue
        rows.append((row_idx, offer, str(raw_article).strip()))
    return rows


def collect_lookup_rows(lookup_ws) -> list[tuple[int, str, str]]:
    rows = []
    for row_idx in range(LOOKUP_HEADER_ROW + 1, lookup_ws.max_row + 1):
        raw_article = lookup_ws[f"{LOOKUP_ARTICLE_COL}{row_idx}"].value
        offer = normalize_article(raw_article)
        if not raw_article or not offer:
            continue
        if offer not in offer_to_sku:
            print(f"WARN lookup row {row_idx}: article {raw_article!r} normalized to {offer!r}, not found in offer_to_sku")
            continue
        rows.append((row_idx, offer, str(raw_article).strip()))
    return rows


def chunked(values: list, size: int):
    for start in range(0, len(values), size):
        yield values[start:start + size]


def safe_number(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return 0.0


def fetch_ovh_liters(offers: list[str]) -> dict[str, float]:
    url = "https://api-seller.ozon.ru/v4/product/info/attributes"
    offer_to_target_sku = {str(offer): str(offer_to_sku[offer]) for offer in offers if offer in offer_to_sku}
    sku_to_offer = {sku: offer for offer, sku in offer_to_target_sku.items()}
    result = {}

    for sku_chunk in chunked(list(sku_to_offer.keys()), 100):
        body = {
            "filter": {"sku": sku_chunk, "visibility": "ALL"},
            "limit": 100,
            "sort_dir": "ASC",
        }
        data = post_with_retry(url, body, timeout=120)
        for item in data.get("result", []) or []:
            sku = str(item.get("sku") or "")
            offer = sku_to_offer.get(sku)
            if not offer:
                continue

            width = safe_number(item.get("width"))
            height = safe_number(item.get("height"))
            depth = safe_number(item.get("depth"))

            for attr in item.get("attributes", []) or []:
                name = str(attr.get("name") or "").lower()
                values = attr.get("values", []) or []
                if not values:
                    continue
                val = safe_number(values[0].get("value"))
                if ("ширина" in name or "width" in name) and not width:
                    width = val
                elif ("высота" in name or "height" in name) and not height:
                    height = val
                elif ("глубина" in name or "depth" in name) and not depth:
                    depth = val

            result[offer] = round((width * height * depth) / 1_000_000, 2) if width and height and depth else 0.0

    return result


def fetch_prices(offers: list[str]) -> dict[str, float]:
    url = "https://api-seller.ozon.ru/v5/product/info/prices"
    result = {offer: 0.0 for offer in offers}
    product_to_offer = {str(v): str(k) for k, v in clean_offer_to_product_id.items()}

    product_ids = [clean_offer_to_product_id[offer] for offer in offers if offer in clean_offer_to_product_id]
    for product_chunk in chunked(product_ids, 100):
        body = {
            "cursor": "",
            "filter": {"product_id": product_chunk, "visibility": "ALL"},
            "limit": 100,
        }
        data = post_with_retry(url, body, timeout=60)
        for item in data.get("items", []) or []:
            offer = product_to_offer.get(str(item.get("product_id")))
            if not offer:
                continue
            price_block = item.get("price", {}) or {}
            result[offer] = safe_number(price_block.get("marketing_seller_price"))

    return result


def fetch_available_cluster_counts(offers: list[str]) -> dict[str, int]:
    url = "https://api-seller.ozon.ru/v1/analytics/stocks"
    wanted_offers = set(offers)
    known_cluster_ids = {int(v) for v in clusters_id.values() if str(v).isdigit()}
    sku_to_offer = {str(sku): offer for offer, sku in offer_to_sku.items() if offer in wanted_offers}
    result = {offer: set() for offer in offers}

    data = post_with_retry(url, {"skus": [int(sku) for sku in sku_to_offer]}, timeout=120)
    for item in data.get("items", []) or []:
        offer = normalize_article(item.get("offer_id"))
        if offer not in wanted_offers:
            sku = str(item.get("sku") or "")
            offer = sku_to_offer.get(sku, offer)
        if offer not in wanted_offers:
            continue

        try:
            cluster_id = int(item.get("cluster_id"))
        except Exception:
            continue
        if cluster_id not in known_cluster_ids:
            continue

        available = int(item.get("available_stock_count", 0) or 0)
        if available > 0:
            result.setdefault(offer, set()).add(cluster_id)

    return {offer: len(cluster_ids) for offer, cluster_ids in result.items()}


def apply_updates(main_ws, lookup_ws, main_rows, lookup_rows, ovh, prices, cluster_counts) -> dict[str, int]:
    updated = {"main_volume": 0, "main_price": 0, "lookup_clusters": 0}

    for row_idx, offer, _raw in main_rows:
        volume = ovh.get(offer, 0.0)
        price = prices.get(offer, 0.0)
        if volume:
            main_ws[f"{MAIN_VOLUME_COL}{row_idx}"] = volume
            updated["main_volume"] += 1
        if price:
            main_ws[f"{MAIN_PRICE_COL}{row_idx}"] = price
            updated["main_price"] += 1

    for row_idx, offer, _raw in lookup_rows:
        lookup_ws[f"{LOOKUP_CLUSTER_COUNT_COL}{row_idx}"] = cluster_counts.get(offer, 0)
        updated["lookup_clusters"] += 1

    return updated


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Update unit economics .xlsx workbook from Ozon API.")
    parser.add_argument("--dry-run", action="store_true", help="Do not upload changed workbook back to Google Drive.")
    parser.add_argument("--file-id", default=UNITHOUR_FILE_ID, help="Google Drive file id to update.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        raise RuntimeError("Отсутствуют OZON_CLIENT_ID / OZON_API_KEY")
    if not args.file_id:
        raise RuntimeError(
            "Не указан ID тестовой копии. Добавь GOOGLE_SHEET_ID_UNITHOUR_TEST в .env "
            "или передай --file-id."
        )

    original_content = download_drive_xlsx(args.file_id)
    workbook = load_workbook(BytesIO(original_content), data_only=False)
    main_ws, lookup_ws = validate_layout(workbook)

    main_rows = collect_main_rows(main_ws)
    lookup_rows = collect_lookup_rows(lookup_ws)
    offers = sorted({offer for _row, offer, _raw in main_rows} | {offer for _row, offer, _raw in lookup_rows})

    print(f"Workbook: sheets ok, main_rows={len(main_rows)}, lookup_rows={len(lookup_rows)}, offers={len(offers)}")
    if not offers:
        print("No offers found, nothing to update.")
        return

    ovh = fetch_ovh_liters(offers)
    prices = fetch_prices(offers)
    cluster_counts = fetch_available_cluster_counts(offers)

    print(f"Fetched: ovh={sum(1 for v in ovh.values() if v)}, prices={sum(1 for v in prices.values() if v)}, clusters={len(cluster_counts)}")
    print("Sample:", [
        {
            "offer": offer,
            "ovh_l": ovh.get(offer),
            "price": prices.get(offer),
            "clusters": cluster_counts.get(offer),
        }
        for offer in offers[:10]
    ])

    updated = apply_updates(main_ws, lookup_ws, main_rows, lookup_rows, ovh, prices, cluster_counts)
    print(f"Prepared updates: {updated}")

    if args.dry_run:
        print("DRY RUN: workbook was not uploaded.")
        return

    backup_path = save_local_backup(original_content)
    output = BytesIO()
    workbook.save(output)
    upload_drive_xlsx(args.file_id, output.getvalue())
    print(f"Uploaded updated workbook. Local backup: {backup_path}")


if __name__ == "__main__":
    main()
