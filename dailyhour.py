_LAST_POSTINGS_CACHE = []
import os
import json
import math
import re
from io import BytesIO
from datetime import datetime, timedelta, timezone
from constants import offer_to_sku, clusters_id
import requests
import time
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from dotenv import load_dotenv
from openpyxl import load_workbook

import finance_accrual

load_dotenv()

# ---- ALERT CONFIG (VK TEAMS) ----
VK_TEAMS_API_BASE = os.getenv("VKTEAMS_API", "").rstrip("/")
VK_TEAMS_BOT_URL = f"{VK_TEAMS_API_BASE}/messages/sendText" if VK_TEAMS_API_BASE else ""
VK_TEAMS_TOKEN = os.getenv("VKTEAMS_BOT_TOKEN", "")
#VK_TEAMS_CHAT_ID = os.getenv("VK_CHAT_ID", "") #чат аналитика МП
VK_TEAMS_CHAT_ID = os.getenv("VK_TEAMS_CHAT_ID", "") #Часовой озон
#VK_TEAMS_CHAT_ID = "user@example.com"
OZON_PERF_CLIENT_ID = os.getenv("OZON_PERF_CLIENT_ID", "")
OZON_PERF_CLIENT_SECRET = os.getenv("OZON_PERF_CLIENT_SECRET", "")

def send_vk_teams_alert(text: str):
    if not VK_TEAMS_BOT_URL or not VK_TEAMS_TOKEN or not VK_TEAMS_CHAT_ID:
        print("VK TEAMS WARNING: alert env vars are not fully configured, skip sending alert")
        print("VK DEBUG:", repr(VK_TEAMS_BOT_URL), repr(VK_TEAMS_TOKEN), repr(VK_TEAMS_CHAT_ID))
        return

    try:
        resp = requests.post(
            VK_TEAMS_BOT_URL,
            data={
                "token": VK_TEAMS_TOKEN,
                "chatId": VK_TEAMS_CHAT_ID,
                "text": text
            },
            timeout=10
        )
        print("VK TEAMS STATUS:", resp.status_code)
        print("VK TEAMS RESPONSE:", resp.text)

        try:
            resp_json = resp.json()
            status = resp_json.get("status", {}) if isinstance(resp_json, dict) else {}
            if status.get("code") not in (None, 0):
                raise RuntimeError(f"VK Teams API error: {resp.text}")
        except ValueError:
            pass

        resp.raise_for_status()
    except Exception as e:
        print("VK TEAMS ALERT ERROR:", e)


def calc_cost(price, ovh_val):
    """
    Replicates simplified cost formula used in Google Sheets.
    Needed to estimate margin in Python for alerts.
    """
    if not price:
        return None

    # commission
    commission = price * (0.20 if price < 300 else 0.39)

    # other fees
    fees = price * 0.015 + price * 0.05

    # logistics (same logic as sheet but simplified)
    if ovh_val < 300:
        logistics = 0
    elif ovh_val <= 1:
        logistics = 46.77
    elif ovh_val <= 2:
        logistics = 56.94
    elif ovh_val <= 3:
        logistics = 67.11
    elif ovh_val <= 190:
        logistics = 67.11 + (ovh_val - 3) * 15.25
    elif ovh_val <= 1000:
        logistics = 67.11 + (190 - 3) * 15.25 + (ovh_val - 190) * 6.1
    else:
        logistics = 7859.86

    return commission + fees + logistics
# ---- TIME HELPERS ----
MSK = timezone(timedelta(hours=3))
NSK = timezone(timedelta(hours=7))  # Новосибирск

# ---- ENV / CONFIG ----
OZON_CLIENT_ID = os.getenv("OZON_CLIENT_ID")
OZON_API_KEY = os.getenv("OZON_API_KEY")
GOOGLE_SHEET_NAME_DAY = os.getenv("GOOGLE_SHEET_NAME_DAY")
WORKSHEET_NAME = os.getenv("WORKSHEET_NAME", "Час")
GOOGLE_SHEET_NAME_MANUAL = os.getenv("GOOGLE_SHEET_NAME_MANUAL")
GOOGLE_SHEET_ID_MANUAL = os.getenv("GOOGLE_SHEET_ID_MANUAL")
GOOGLE_SHEET_WORKSHEET_MANUAL = os.getenv("GOOGLE_SHEET_WORKSHEET_MANUAL", "Калькулятор прибыли")
GOOGLE_SHEET_ID_HOURLY_CALC = os.getenv("GOOGLE_SHEET_ID_HOURLY_CALC")
GOOGLE_SHEET_NAME_HOURLY_CALC = os.getenv("GOOGLE_SHEET_NAME_HOURLY_CALC")
GOOGLE_SHEET_WORKSHEET_HOURLY_CALC = os.getenv("GOOGLE_SHEET_WORKSHEET_HOURLY_CALC", "Калькулятор прибыли")
# same as in other scripts: GOOGLE_CREDS points to creds.json path
GOOGLE_CREDS_PATH = os.getenv("GOOGLE_CREDS") or os.getenv("GOOGLE_CREDS_PATH") or "creds.json"
DAYS_BACK = int(os.getenv("DAYS_BACK", "1"))  # 1=yesterday, 2=day before yesterday
STATE_FILE = os.path.join(os.path.dirname(__file__), "dailyhour_state.json")

_missing = [k for k, v in {
    "OZON_CLIENT_ID": OZON_CLIENT_ID,
    "OZON_API_KEY": OZON_API_KEY,
    "GOOGLE_SHEET_NAME_DAY": GOOGLE_SHEET_NAME_DAY,
}.items() if not v]

if _missing:
    raise RuntimeError(
        "Missing required environment variables: " + ", ".join(_missing)
    )

HEADERS = {
    "Client-Id": OZON_CLIENT_ID,
    "Api-Key": OZON_API_KEY,
    "Content-Type": "application/json",
}

# ----------------------------------------------------
# HTTP RETRY (survive 5xx / timeouts)
# ----------------------------------------------------

def post_with_retry(url, headers, body, timeout=60):
    attempt = 0
    while True:
        try:
            resp = requests.post(url, headers=headers, json=body, timeout=timeout)
            if resp.status_code == 429 or resp.status_code >= 500:
                raise requests.exceptions.HTTPError(f"Server error {resp.status_code}", response=resp)
            return resp
        except Exception as e:
            attempt += 1
            wait = min(30 * attempt, 300)
            print(f"RETRY {attempt}: {url} -> {e}. Sleep {wait}s")
            time.sleep(wait)


def get_with_retry(url, headers=None, params=None, timeout=60):
    attempt = 0
    while True:
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=timeout)
            if resp.status_code >= 500:
                raise requests.exceptions.HTTPError(f"Server error {resp.status_code}", response=resp)
            return resp
        except Exception as e:
            attempt += 1
            wait = min(30 * attempt, 300)
            print(f"RETRY {attempt}: {url} -> {e}. Sleep {wait}s")
            time.sleep(wait)

# ----------------------------------------------------
# GOOGLE SHEETS
# ----------------------------------------------------

def connect_sheet():
    client = get_gspread_client()
    return client.open(GOOGLE_SHEET_NAME_DAY).worksheet(WORKSHEET_NAME)


def get_google_creds():
    scope = ["https://spreadsheets.google.com/feeds",
             "https://www.googleapis.com/auth/drive"]

    return ServiceAccountCredentials.from_json_keyfile_name(GOOGLE_CREDS_PATH, scope)


def get_gspread_client():
    creds = get_google_creds()
    return gspread.authorize(creds)


def is_unsupported_document_error(exc):
    error_data = getattr(exc, "args", [{}])[0]
    if isinstance(error_data, dict):
        message = str(error_data.get("message", ""))
        return "This operation is not supported for this document" in message
    return False


def download_drive_file_content(file_id, log_prefix):
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"

    try:
        client = get_gspread_client()
        resp = client.http_client.request("get", url)
        resp.raise_for_status()
        print(f"{log_prefix} DEBUG: downloaded via gspread http client")
        return resp.content
    except Exception as e:
        print(f"{log_prefix} WARNING: gspread http download failed -> {e}")

    creds = get_google_creds()
    token = creds.get_access_token().access_token
    headers = {"Authorization": f"Bearer {token}"}

    download_attempts = [
        ("requests default session", None),
        ("requests session trust_env=False", False),
    ]

    last_error = None
    for label, trust_env in download_attempts:
        try:
            session = requests.Session()
            if trust_env is not None:
                session.trust_env = trust_env
            resp = session.get(url, headers=headers, timeout=60)
            resp.raise_for_status()
            print(f"{log_prefix} DEBUG: downloaded via {label}")
            return resp.content
        except Exception as e:
            last_error = e
            print(f"{log_prefix} WARNING: {label} failed -> {e}")

    raise RuntimeError(f"{log_prefix} ERROR: failed to download file from Drive -> {last_error}")


def read_manual_xlsx_rows_by_id(file_id, worksheet_name):
    content = download_drive_file_content(file_id, "MANUAL SHEET")

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if worksheet_name not in workbook.sheetnames:
        raise RuntimeError(
            f"MANUAL SHEET ERROR: worksheet {worksheet_name!r} not found in xlsx file"
        )

    ws = workbook[worksheet_name]
    rows = []
    for row in ws.iter_rows(min_col=1, max_col=77, values_only=True):
        rows.append(list(row))
    return rows


def read_hourly_calc_xlsx_rows_by_id(file_id, worksheet_name):
    content = download_drive_file_content(file_id, "HOURLY CALC")

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if worksheet_name not in workbook.sheetnames:
        raise RuntimeError(
            f"HOURLY CALC ERROR: worksheet {worksheet_name!r} not found in xlsx file"
        )

    ws = workbook[worksheet_name]
    rows = []
    max_cols = 200
    for row in ws.iter_rows(min_col=1, max_col=max_cols, values_only=True):
        rows.append(list(row))
    return rows


def connect_manual_sheet():
    if not GOOGLE_SHEET_ID_MANUAL and not GOOGLE_SHEET_NAME_MANUAL:
        print("MANUAL SHEET WARNING: neither GOOGLE_SHEET_ID_MANUAL nor GOOGLE_SHEET_NAME_MANUAL is set, manual metrics disabled")
        return None

    client = get_gspread_client()
    spreadsheet = None

    try:
        if GOOGLE_SHEET_ID_MANUAL:
            print(f"MANUAL SHEET DEBUG: opening by key {GOOGLE_SHEET_ID_MANUAL}")
            try:
                spreadsheet = client.open_by_key(GOOGLE_SHEET_ID_MANUAL)
            except gspread.exceptions.APIError as e:
                if is_unsupported_document_error(e):
                    raise
                else:
                    raise

        if spreadsheet is None and GOOGLE_SHEET_NAME_MANUAL and not GOOGLE_SHEET_ID_MANUAL:
            print(f"MANUAL SHEET DEBUG: opening by name {GOOGLE_SHEET_NAME_MANUAL}")
            spreadsheet = client.open(GOOGLE_SHEET_NAME_MANUAL)

        print(f"MANUAL SHEET DEBUG: opened spreadsheet '{spreadsheet.title}'")
        return spreadsheet.worksheet(GOOGLE_SHEET_WORKSHEET_MANUAL)

    except gspread.exceptions.SpreadsheetNotFound:
        print(
            "MANUAL SHEET ERROR: spreadsheet not found. "
            f"GOOGLE_SHEET_ID_MANUAL={GOOGLE_SHEET_ID_MANUAL!r}, "
            f"GOOGLE_SHEET_NAME_MANUAL={GOOGLE_SHEET_NAME_MANUAL!r}, "
            f"worksheet={GOOGLE_SHEET_WORKSHEET_MANUAL!r}, "
            f"creds_path={GOOGLE_CREDS_PATH!r}"
        )
        raise
    except gspread.exceptions.WorksheetNotFound:
        print(
            "MANUAL SHEET ERROR: worksheet not found. "
            f"worksheet={GOOGLE_SHEET_WORKSHEET_MANUAL!r}"
        )
        raise
    except gspread.exceptions.APIError as e:
        print(
            "MANUAL SHEET ERROR: Google API rejected the manual document. "
            "Most likely GOOGLE_SHEET_ID_MANUAL points to a non-Google-Sheets file "
            "(for example an .xlsx in Drive) or a wrong document ID."
        )
        raise


def normalize_manual_offer(raw_value):
    if raw_value in (None, ""):
        return ""

    if isinstance(raw_value, (int, float)):
        num = float(raw_value)
        if num.is_integer():
            value = str(int(num))
        else:
            value = str(raw_value)
    else:
        value = str(raw_value).strip()

    if not value:
        return ""

    value = value.replace("\xa0", " ").strip()
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"(?i)\s*ИР\s*$", "", value).strip()
    value = re.sub(r"(?<=\d)\.0+$", "", value)
    if not value:
        return ""

    match = re.match(r"^\s*(\d+)(?:/\d+)?(?:\s*([A-Za-zА-Яа-яЁё]+))?\s*$", value)
    if not match:
        return ""

    normalized = match.group(1)
    suffix = match.group(2)
    if suffix:
        normalized = f"{normalized} {suffix.upper()}"

    return normalized.strip()


def build_offer_lookup_candidates(raw_value):
    if raw_value in (None, ""):
        return []

    if isinstance(raw_value, (int, float)):
        num = float(raw_value)
        if num.is_integer():
            value = str(int(num))
        else:
            value = str(raw_value)
    else:
        value = str(raw_value)

    value = value.replace("\xa0", " ").strip().upper()
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"(?i)\s*ИР\s*$", "", value).strip()
    if not value:
        return []

    candidates = []

    def _add(candidate):
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    _add(normalize_manual_offer(value))

    slash_removed = re.sub(r"(?<=\d)/\d+", "", value).strip()
    _add(normalize_manual_offer(slash_removed))

    base_match = re.match(r"^\s*(\d+)", slash_removed)
    if base_match:
        _add(base_match.group(1))

    return candidates


def resolve_offer_key(raw_value, offer_map):
    for candidate in build_offer_lookup_candidates(raw_value):
        if candidate in offer_map:
            return candidate
    return ""


def normalize_margin_percent(value):
    if value in (None, ""):
        return ""

    if isinstance(value, (int, float)):
        num = float(value)
        # If Excel/Google gives a true percentage value like 0.23, convert it to 23
        if -1 <= num <= 1 and num != 0:
            return num * 100
        return num

    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return ""

        has_percent_sign = "%" in raw
        raw = raw.replace(" ", "").replace(",", ".").replace("%", "")
        if not raw:
            return ""

        try:
            num = float(raw)
        except Exception:
            return value

        # "23%" -> 23, "23" -> 23, "0.23" -> 23
        if has_percent_sign:
            return num
        if -1 <= num <= 1 and num != 0:
            return num * 100
        return num

    return value


def load_manual_metrics(offer_map):
    try:
        sheet = connect_manual_sheet()
        if sheet is None:
            return {}
        rows = sheet.get("A:BY", value_render_option="UNFORMATTED_VALUE")
        print("MANUAL SHEET DEBUG: source=gspread")
    except gspread.exceptions.APIError as e:
        if GOOGLE_SHEET_ID_MANUAL and is_unsupported_document_error(e):
            print("MANUAL SHEET WARNING: source document is xlsx, reading via Drive API + openpyxl")
            rows = read_manual_xlsx_rows_by_id(GOOGLE_SHEET_ID_MANUAL, GOOGLE_SHEET_WORKSHEET_MANUAL)
            print("MANUAL SHEET DEBUG: source=xlsx")
        else:
            raise

    metrics = {}
    duplicates = 0
    valid_matches = 0
    skipped_invalid_offer = 0
    skipped_missing_offer = 0
    skipped_invalid_examples = []
    skipped_missing_examples = []

    for idx, row in enumerate(rows, start=1):
        raw_offer = row[0] if len(row) > 0 else ""
        normalized_offer = resolve_offer_key(raw_offer, offer_map)
        if not normalized_offer:
            skipped_invalid_offer += 1
            if len(skipped_invalid_examples) < 5 and raw_offer not in (None, ""):
                skipped_invalid_examples.append(str(raw_offer))
            continue
        if normalized_offer not in offer_map:
            skipped_missing_offer += 1
            if len(skipped_missing_examples) < 5:
                skipped_missing_examples.append(normalized_offer)
            continue

        cost_share = row[75] if len(row) > 75 else ""
        margin_percent = row[76] if len(row) > 76 else ""

        cost_share = "" if cost_share in (None, "") else cost_share
        margin_percent = normalize_margin_percent(margin_percent)
        is_non_empty = cost_share != "" or margin_percent != ""

        if normalized_offer in metrics:
            duplicates += 1
            print(f"MANUAL SHEET WARNING: duplicate offer {normalized_offer} at row {idx}")

        if not is_non_empty:
            continue

        if normalized_offer not in metrics:
            valid_matches += 1
        metrics[normalized_offer] = {
            "cost_share": cost_share,
            "margin_percent": margin_percent,
        }

    print(
        "MANUAL SHEET:",
        f"rows_read={len(rows)}",
        f"valid_offers={valid_matches}",
        f"duplicates={duplicates}",
        f"invalid_offer_rows={skipped_invalid_offer}",
        f"missing_in_offer_to_sku={skipped_missing_offer}",
    )
    if skipped_invalid_examples:
        print(f"MANUAL SHEET DEBUG: invalid_offer_examples={skipped_invalid_examples}")
    if skipped_missing_examples:
        print(f"MANUAL SHEET DEBUG: missing_offer_examples={skipped_missing_examples}")
    return metrics


def connect_hourly_calc_sheet():
    if not GOOGLE_SHEET_ID_HOURLY_CALC and not GOOGLE_SHEET_NAME_HOURLY_CALC:
        raise RuntimeError(
            "HOURLY CALC ERROR: set GOOGLE_SHEET_ID_HOURLY_CALC or GOOGLE_SHEET_NAME_HOURLY_CALC"
        )

    client = get_gspread_client()
    spreadsheet = None

    try:
        if GOOGLE_SHEET_ID_HOURLY_CALC:
            print(f"HOURLY CALC DEBUG: opening by key {GOOGLE_SHEET_ID_HOURLY_CALC}")
            spreadsheet = client.open_by_key(GOOGLE_SHEET_ID_HOURLY_CALC)
        elif GOOGLE_SHEET_NAME_HOURLY_CALC:
            print(f"HOURLY CALC DEBUG: opening by name {GOOGLE_SHEET_NAME_HOURLY_CALC}")
            spreadsheet = client.open(GOOGLE_SHEET_NAME_HOURLY_CALC)

        print(f"HOURLY CALC DEBUG: opened spreadsheet '{spreadsheet.title}'")
        return spreadsheet.worksheet(GOOGLE_SHEET_WORKSHEET_HOURLY_CALC)

    except gspread.exceptions.SpreadsheetNotFound:
        print(
            "HOURLY CALC ERROR: spreadsheet not found. "
            f"GOOGLE_SHEET_ID_HOURLY_CALC={GOOGLE_SHEET_ID_HOURLY_CALC!r}, "
            f"GOOGLE_SHEET_NAME_HOURLY_CALC={GOOGLE_SHEET_NAME_HOURLY_CALC!r}, "
            f"worksheet={GOOGLE_SHEET_WORKSHEET_HOURLY_CALC!r}, "
            f"creds_path={GOOGLE_CREDS_PATH!r}"
        )
        raise
    except gspread.exceptions.WorksheetNotFound:
        print(
            "HOURLY CALC ERROR: worksheet not found. "
            f"worksheet={GOOGLE_SHEET_WORKSHEET_HOURLY_CALC!r}"
        )
        raise
    except gspread.exceptions.APIError as e:
        print(
            "HOURLY CALC ERROR: Google API rejected the calculator document. "
            "Most likely GOOGLE_SHEET_ID_HOURLY_CALC points to a non-Google-Sheets file "
            "(for example an .xlsx in Drive) or a wrong document ID."
        )
        raise


def normalize_header_name(value):
    if value in (None, ""):
        return ""
    text = str(value).replace("\n", " ").replace("\xa0", " ").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def parse_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace("\xa0", "").replace(" ", "").replace("%", "").replace(",", ".")
    try:
        return float(raw)
    except Exception:
        return None


def find_header_row_and_columns(rows):
    header_aliases = {
        "article": {"артикул"},
        "margin_rub": {"маржа fbo, руб"},
        "margin_pct": {"маржа fbo, %"},
        "current_price": {"текущая цена, руб", "текущая цена руб"},
    }

    for row_idx, row in enumerate(rows):
        normalized = [normalize_header_name(cell) for cell in row]
        if not normalized:
            continue

        found = {}
        for key, aliases in header_aliases.items():
            for col_idx, cell in enumerate(normalized):
                if cell in aliases:
                    found[key] = col_idx
                    break

        if "article" in found and "margin_rub" in found and "margin_pct" in found:
            print(
                "HOURLY CALC DEBUG:",
                f"header_row={row_idx + 1}",
                f"article_col={found['article'] + 1}",
                f"margin_rub_col={found['margin_rub'] + 1}",
                f"margin_pct_col={found['margin_pct'] + 1}",
                f"current_price_col={(found.get('current_price', -1) + 1) if 'current_price' in found else 'n/a'}",
            )
            return row_idx, found

    raise RuntimeError(
        "HOURLY CALC ERROR: required headers not found. "
        "Need at least 'Артикул', 'Маржа FBO, руб', 'Маржа FBO, %'."
    )


def load_hourly_economics_from_calc(offer_map):
    try:
        sheet = connect_hourly_calc_sheet()
        rows = sheet.get_all_values()
        print("HOURLY CALC DEBUG: source=gspread")
    except gspread.exceptions.APIError as e:
        if GOOGLE_SHEET_ID_HOURLY_CALC and is_unsupported_document_error(e):
            print("HOURLY CALC WARNING: source document is xlsx, reading via Drive API + openpyxl")
            rows = read_hourly_calc_xlsx_rows_by_id(
                GOOGLE_SHEET_ID_HOURLY_CALC,
                GOOGLE_SHEET_WORKSHEET_HOURLY_CALC,
            )
            print("HOURLY CALC DEBUG: source=xlsx")
        else:
            raise RuntimeError(f"HOURLY CALC ERROR: failed to read calculator sheet -> {e}") from e
    except Exception as e:
        raise RuntimeError(f"HOURLY CALC ERROR: failed to read calculator sheet -> {e}") from e

    if not rows:
        raise RuntimeError("HOURLY CALC ERROR: calculator worksheet is empty")

    header_row_idx, columns = find_header_row_and_columns(rows)
    economics = {}
    duplicates = 0
    skipped_invalid_offer = 0
    skipped_missing_offer = 0
    invalid_examples = []
    missing_examples = []

    for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        raw_article = row[columns["article"]] if len(row) > columns["article"] else ""
        offer = resolve_offer_key(raw_article, offer_map)
        if not raw_article:
            continue
        if not offer:
            skipped_missing_offer += 1
            if len(missing_examples) < 5:
                missing_examples.append(str(raw_article))
            print(f"HOURLY CALC WARNING: row={row_idx} article={raw_article!r} not matched to offer")
            continue

        unit_margin_rub = parse_number(row[columns["margin_rub"]]) if len(row) > columns["margin_rub"] else None
        unit_margin_pct = parse_number(row[columns["margin_pct"]]) if len(row) > columns["margin_pct"] else None
        current_price = parse_number(row[columns["current_price"]]) if "current_price" in columns and len(row) > columns["current_price"] else None

        if unit_margin_rub is None and unit_margin_pct is None and current_price is None:
            skipped_invalid_offer += 1
            if len(invalid_examples) < 5:
                invalid_examples.append(str(raw_article))
            continue

        if offer in economics:
            duplicates += 1
            print(f"HOURLY CALC WARNING: duplicate offer {offer} at row {row_idx}, overwrite with latest row")

        economics[offer] = {
            "raw_article": str(raw_article).strip(),
            "unit_margin_rub": unit_margin_rub,
            "unit_margin_pct": unit_margin_pct,
            "current_price": current_price,
            "source_row": row_idx,
        }

    print(
        "HOURLY CALC:",
        f"rows_read={len(rows)}",
        f"offers_loaded={len(economics)}",
        f"duplicates={duplicates}",
        f"rows_without_metrics={skipped_invalid_offer}",
        f"rows_unmatched={skipped_missing_offer}",
    )
    if invalid_examples:
        print(f"HOURLY CALC DEBUG: empty_metric_examples={invalid_examples}")
    if missing_examples:
        print(f"HOURLY CALC DEBUG: unmatched_examples={missing_examples}")

    return economics


def read_table_structure(sheet):
    # начиная с 3 строки (индекс 2)
    rows = sheet.get_all_values()[2:]
    out = []
    for r in rows:
        if not r or not r[0]:
            continue
        offer = str(r[0]).strip()
        metric = str(r[1]).strip() if len(r) > 1 else ""
        out.append((offer, metric))
    return out


def get_or_create_date_column(sheet, date_label, hour_label):
    header_dates = sheet.row_values(1)
    header_hours = sheet.row_values(2)

    max_len = max(len(header_dates), len(header_hours))

    for i in range(2, max_len):  # начиная с колонки C
        d = header_dates[i] if i < len(header_dates) else ""
        h = header_hours[i] if i < len(header_hours) else ""
        if d == date_label and h == hour_label:
            return i + 1

    col = max(max_len, 2) + 1
    if col < 3:
        col = 3

    sheet.update_cell(1, col, date_label)
    sheet.update_cell(2, col, hour_label)
    return col


def write_values(sheet, values, col):
    # очищаем старые данные в колонке даты (с 3 строки до конца листа)
    col_letter = gspread.utils.rowcol_to_a1(1, col).rstrip('1')
    max_rows = sheet.row_count
    clear_range = f"{col_letter}3:{col_letter}{max_rows}"
    sheet.update(values=[[""]]* (max_rows-2), range_name=clear_range)

    # пишем новые данные
    end_row = len(values) + 2
    rng = gspread.utils.rowcol_to_a1(3, col) + ":" + gspread.utils.rowcol_to_a1(end_row, col)
    sheet.update(values=[[v] for v in values], range_name=rng, value_input_option="USER_ENTERED")

# ----------------------------------------------------
# OZON API — ПОСТИНГИ (ЦЕНА)
# ----------------------------------------------------

def get_postings(date_from, date_to):

    url = "https://api-seller.ozon.ru/v3/posting/fbo/list"

    postings = []
    cursor = ""

    while True:
        body = {
            "sort_dir": "ASC",
            "filter": {
                "since": date_from,
                "to": date_to,
            },
            "limit": 100,
            "cursor": cursor,
            "translit": True,
            "with": {
                "analytics_data": True,
                "financial_data": True,
                "legal_info": False
            }
        }

        resp = post_with_retry(url, HEADERS, body, timeout=60)
        resp.raise_for_status()
        r = resp.json()

        batch = r.get("postings", [])
        if not batch:
            break

        postings.extend(batch)
        cursor = r.get("cursor", "")
        if not r.get("has_next", False):
            break

    return postings


def get_finance_totals(date):
    # МИГРИРОВАНО с устаревающего /v3/finance/transaction/totals на
    # finance/accrual: свод начислений за день (by-day).
    # Для текущего дня by-day отдаёт накопительные начисления «на момент
    # запроса» — начислений из будущего не бывает, поэтому это эквивалент
    # старого окна 00:00→сейчас (сверено 2026-06-19). Те же ключи и знаки.
    return finance_accrual.totals(date)


def calc_commission_from_totals(totals):
    def _to_float(value):
        try:
            if isinstance(value, bool):
                return 0.0
            return float(value)
        except Exception:
            return 0.0

    parts = {
        "sale_commission": _to_float(totals.get("sale_commission", 0)),
        "processing_and_delivery": _to_float(totals.get("processing_and_delivery", 0)),
        "refunds_and_cancellations": _to_float(totals.get("refunds_and_cancellations", 0)),
        "services_amount": _to_float(totals.get("services_amount", 0)),
        "compensation_amount": _to_float(totals.get("compensation_amount", 0)),
        "money_transfer": _to_float(totals.get("money_transfer", 0)),
        "others_amount": _to_float(totals.get("others_amount", 0)),
    }

    losses = 0.0
    for key in (
        "sale_commission",
        "processing_and_delivery",
        "services_amount",
        "compensation_amount",
        "money_transfer",
        "others_amount",
    ):
        value = parts[key]
        if value < 0:
            losses += abs(value)

    losses += abs(parts["refunds_and_cancellations"])

    print(f"HOURLY TOTALS COMPONENTS: {parts}")
    print(f"HOURLY TOTALS COMMISSION: {round(losses, 2)}")
    return losses


def get_performance_access_token():
    if not OZON_PERF_CLIENT_ID or not OZON_PERF_CLIENT_SECRET:
        print("PERF WARNING: OZON_PERF_CLIENT_ID / OZON_PERF_CLIENT_SECRET are not configured")
        return None

    url = "https://api-performance.ozon.ru/api/client/token"
    body = {
        "client_id": OZON_PERF_CLIENT_ID,
        "client_secret": OZON_PERF_CLIENT_SECRET,
        "grant_type": "client_credentials",
    }

    try:
        resp = requests.post(url, data=body, timeout=30)
        resp.raise_for_status()
        token_data = resp.json() or {}
        return token_data.get("access_token")
    except Exception as e:
        print(f"PERF WARNING: failed to get token -> {e}")
        return None


def _safe_number(value) -> float:
    try:
        if isinstance(value, bool):
            return 0.0
        if isinstance(value, str):
            value = value.replace(" ", "").replace(",", ".")
        return float(value or 0)
    except Exception:
        return 0.0


def fetch_ad_spend(date_str):
    token = get_performance_access_token()
    if not token:
        return 0.0

    url = "https://api-performance.ozon.ru/api/client/statistics/expense/json"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    params = {
        "dateFrom": date_str,
        "dateTo": date_str,
    }

    try:
        response = get_with_retry(url, headers=headers, params=params, timeout=30)
        response.raise_for_status()
        data = response.json() or {}
        print(f"HOURLY PERF EXPENSE RAW: {data}")
        rows = data.get("rows", []) or []
        total_spend = 0.0
        seen_ids = set()

        for row in rows:
            row_id = row.get("id")
            title = str(row.get("title", "")).strip()

            if row_id in seen_ids:
                continue
            seen_ids.add(row_id)

            if "пакет" in title.lower() or "спецпроект" in title.lower() or not any(ch.isdigit() for ch in title):
                continue

            total_spend += _safe_number(row.get("moneySpent", 0))

        print(f"HOURLY PERF EXPENSE TOTAL: date={date_str} spend={round(total_spend, 2)}")
        return total_spend
    except Exception as e:
        print(f"PERF WARNING: failed to fetch ad spend -> {e}")
        return 0.0


def build_hour_summary_metrics(sales_stats, hourly_economics, finance_totals):
    total_orders = 0.0
    total_margin_rub = 0.0
    margin_offers = 0
    missing_margin_offers = []

    for offer, sale in sales_stats.items():
        qty = sale.get("qty", 0) or 0
        revenue = sale.get("revenue", 0.0) or 0.0
        total_orders += revenue

        economics = hourly_economics.get(offer) or {}
        unit_margin_rub = economics.get("unit_margin_rub")
        if qty > 0 and revenue > 0 and unit_margin_rub is not None:
            total_margin_rub += qty * unit_margin_rub
            margin_offers += 1
        elif qty > 0 and revenue > 0:
            if len(missing_margin_offers) < 10:
                missing_margin_offers.append(offer)

    total_commission = calc_commission_from_totals(finance_totals)
    total_margin_pct = (total_margin_rub / total_orders * 100) if total_orders > 0 else 0.0

    if missing_margin_offers:
        print(f"HOURLY SUMMARY WARNING: missing unit_margin_rub for offers={missing_margin_offers}")

    print(
        "HOURLY CHAT SUMMARY:",
        f"orders={round(total_orders, 2)}",
        f"margin_rub={round(total_margin_rub, 2)}",
        f"margin_pct={round(total_margin_pct, 2)}",
        f"commission={round(total_commission, 2)}",
        f"sales_offers={len(sales_stats)}",
        f"margin_offers={margin_offers}",
        f"missing_margin_offers={len(missing_margin_offers)}",
    )

    return {
        "orders": total_orders,
        "margin_rub": total_margin_rub,
        "margin_pct": total_margin_pct,
        "commission": total_commission,
        "sales_offers": len(sales_stats),
        "margin_offers": margin_offers,
        "missing_margin_offers": len(missing_margin_offers),
    }


def format_money(value):
    try:
        return f"{float(value):,.0f} ₽"
    except Exception:
        return "н/д"


def format_percent(value):
    try:
        return f"{float(value):.1f} %"
    except Exception:
        return "0.0 %"


def safe_round(value, digits=2):
    try:
        return round(float(value), digits)
    except Exception:
        return None


def to_ozon_utc_iso(dt_value):
    return dt_value.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")


def calc_delta_value(current_value, previous_value):
    delta = float(current_value or 0) - float(previous_value or 0)
    if delta < 0:
        print(
            "HOURLY SUMMARY WARNING:",
            f"negative delta current={current_value} previous={previous_value}, clamp to 0"
        )
        return 0.0
    return delta


def load_hourly_state():
    if not os.path.exists(STATE_FILE):
        return None
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"STATE WARNING: failed to read {STATE_FILE} -> {e}")
        return None


def save_hourly_state(state):
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
        print(f"STATE DEBUG: saved {STATE_FILE}")
    except Exception as e:
        print(f"STATE WARNING: failed to save {STATE_FILE} -> {e}")


def build_hour_summary_from_state(current_totals, prev_state, state_date, hour_label, hour_end_msk):
    current_orders = float(current_totals.get("accruals_for_sale", 0) or 0)
    current_finance_commission = calc_commission_from_totals(current_totals)
    current_ad_spend = fetch_ad_spend(state_date)
    current_commission = current_finance_commission + current_ad_spend

    prev_orders = 0.0
    prev_commission = 0.0
    prev_ad_spend = 0.0
    has_same_day_state = bool(prev_state and prev_state.get("date_msk") == state_date)
    if has_same_day_state:
        prev_orders = float(prev_state.get("orders_cumulative", 0) or 0)
        prev_commission = float(prev_state.get("commission_cumulative", 0) or 0)
        prev_ad_spend = float(prev_state.get("ad_spend_cumulative", 0) or 0)

    if has_same_day_state:
        orders_hour = calc_delta_value(current_orders, prev_orders)
        commission_hour = calc_delta_value(current_commission, prev_commission)
    else:
        orders_hour = 0.0
        commission_hour = 0.0

    ad_spend_hour = calc_delta_value(current_ad_spend, prev_ad_spend) if has_same_day_state else 0.0

    print(
        "HOURLY STATE SUMMARY:",
        f"date_msk={state_date}",
        f"same_day_state={has_same_day_state}",
        f"current_orders={safe_round(current_orders, 2)}",
        f"prev_orders={safe_round(prev_orders, 2)}",
        f"orders_hour={safe_round(orders_hour, 2)}",
        f"current_finance_commission={safe_round(current_finance_commission, 2)}",
        f"current_ad_spend={safe_round(current_ad_spend, 2)}",
        f"current_commission={safe_round(current_commission, 2)}",
        f"prev_commission={safe_round(prev_commission, 2)}",
        f"commission_hour={safe_round(commission_hour, 2)}",
        f"prev_ad_spend={safe_round(prev_ad_spend, 2)}",
        f"ad_spend_hour={safe_round(ad_spend_hour, 2)}",
    )

    return {
        "orders": orders_hour,
        "commission": commission_hour,
        "finance_commission": calc_delta_value(current_finance_commission, float(prev_state.get("finance_commission_cumulative", 0) or 0)) if has_same_day_state else 0.0,
        "ad_spend": ad_spend_hour,
        "state": {
            "date_msk": state_date,
            "hour_label": hour_label,
            "hour_end_msk": hour_end_msk.isoformat(),
            "orders_cumulative": current_orders,
            "commission_cumulative": current_commission,
            "finance_commission_cumulative": current_finance_commission,
            "ad_spend_cumulative": current_ad_spend,
        },
    }


def build_hourly_vk_message(hour_label, hour_summary, loss_alerts, low_margin_alerts):
    cost_share_pct = hour_summary.get("cost_share_pct")
    parts = [
        hour_label,
        "",
        f"Сумма заказов: {format_money(hour_summary.get('orders', 0))}",
        f"Маржа: {format_money(hour_summary.get('margin_rub', 0))}",
        f"Маржинальность: {format_percent(hour_summary.get('margin_pct', 0))}",
        f"Комиссия: {format_money(hour_summary.get('commission', 0))}",
        f"Доля затрат: {format_percent(cost_share_pct) if cost_share_pct is not None else 'н/д'}",
    ]

    alert_parts = []
    if loss_alerts:
        alert_parts.append("УБЫТОК (<0%)\n" + "\n".join(loss_alerts))
    if low_margin_alerts:
        alert_parts.append("НИЗКАЯ МАРЖА (<10%)\n" + "\n".join(low_margin_alerts))

    if alert_parts:
        parts.extend(["", ""])
        parts.append("\n\n".join(alert_parts))

    return "\n".join(parts)


def build_sku_to_offer_index():
    const_sku_to_offer = {}
    for offer_id, sku_val in offer_to_sku.items():
        try:
            sku_int = int(sku_val)
            const_sku_to_offer[sku_int] = str(offer_id)
            const_sku_to_offer[str(sku_int)] = str(offer_id)
        except Exception:
            const_sku_to_offer[str(sku_val)] = str(offer_id)
    return const_sku_to_offer


def extract_hour_sales_stats(postings):

    stats = {}

    # build normalized reverse map: sku(int/str) -> our offer_id
    const_sku_to_offer = build_sku_to_offer_index()

    def _map_offer(sku):
        if sku in const_sku_to_offer:
            return const_sku_to_offer[sku]
        try:
            return const_sku_to_offer.get(int(sku))
        except Exception:
            pass
        return const_sku_to_offer.get(str(sku))

    def _num(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    def _qty(v):
        try:
            q = int(v)
            return q if q > 0 else 0
        except Exception:
            return 0

    def _add_sale(offer, price, qty):
        if not offer or price <= 0 or qty <= 0:
            return
        item = stats.setdefault(offer, {"qty": 0, "revenue": 0.0, "avg_price": 0.0})
        item["qty"] += qty
        item["revenue"] += price * qty

    for posting in postings:

        used_financial = set()

        # ---- 1. try real financial price ----
        for fin in posting.get("financial_data", {}).get("products", []) or []:
            sku = fin.get("product_id")
            price = _num(fin.get("price"))
            qty = _qty(fin.get("quantity"))

            offer = _map_offer(sku)
            if not offer or price <= 0 or qty <= 0:
                continue

            _add_sale(offer, price, qty)
            used_financial.add(offer)

        # ---- 2. fallback to products price (Ozon sometimes returns quantity=0 in financial_data) ----
        for item in posting.get("products", []):
            sku = item.get("sku")
            offer = _map_offer(sku)
            if not offer or offer in used_financial:
                continue

            price_raw = item.get("price")
            price = _num(price_raw.get("amount", 0) if isinstance(price_raw, dict) else price_raw)
            qty = _qty(item.get("quantity"))
            if price > 0:
                _add_sale(offer, price, qty)

    for offer, item in stats.items():
        qty = item.get("qty", 0)
        revenue = item.get("revenue", 0.0)
        item["avg_price"] = revenue / qty if qty > 0 else 0.0

    return stats


def extract_real_prices(postings):
    sales_stats = extract_hour_sales_stats(postings)
    return {
        offer: item["avg_price"]
        for offer, item in sales_stats.items()
        if item.get("qty", 0) > 0
    }

# ----------------------------------------------------
# OZON API — ГАБАРИТЫ (ОВХ)
# ----------------------------------------------------

def get_dimensions(offers):

    url = "https://api-seller.ozon.ru/v4/product/info/attributes"
    result = {}

    # build offer->sku and sku->offer maps using constants
    offer_list = [str(o) for o in offers]
    sku_map = {str(k): v for k, v in offer_to_sku.items() if str(k) in offer_list}
    sku_to_offer = {v: k for k, v in sku_map.items()}
    skus = list(sku_to_offer.keys())

    for i in range(0, len(skus), 100):
        chunk = skus[i:i+100]
        body = {
            "filter": {
                "sku": chunk,
                "visibility": "ALL"
            },
            "limit": 100,
            "sort_dir": "ASC"
        }
        resp = post_with_retry(url, HEADERS, body, timeout=60)
        resp.raise_for_status()
        r = resp.json()

        for item in r.get("result", []):
            sku = item.get("sku")
            offer = sku_to_offer.get(sku)
            w = item.get("width", 0)
            h = item.get("height", 0)
            d = item.get("depth", 0)

            ovh = (w*h*d)/1_000_000 if w and h and d else 0
            if offer:
                result[str(offer)] = ovh

    return result

# OZON API — SVD PER SKU
# ----------------------------------------------------

def get_svd_per_offer(offers):
    # СВД считаем по /details как в ЛК:
    # 1) берём готовое metrics.average_delivery_time,
    # 2) считаем по всем кластерам доставки,
    # 3) агрегируем по артикулу взвешенно по orders_count.total.
    offer_list = [str(o) for o in offers]
    offer_to_target_sku = {}
    for offer in offer_list:
        sku = offer_to_sku.get(offer)
        if sku is None:
            continue
        try:
            offer_to_target_sku[offer] = int(sku)
        except Exception:
            continue
    if not offer_to_target_sku:
        print("SVD DEBUG: no mapped offers -> empty SVD")
        return {}

    target_skus = set(offer_to_target_sku.values())
    sku_to_offer = {v: k for k, v in offer_to_target_sku.items()}
    target_offers = set(offer_to_target_sku.keys())

    used_clusters = {int(v) for v in clusters_id.values() if str(v).isdigit()}
    svd_periods = ["FOUR_WEEKS", "EIGHT_WEEKS"]

    print("SVD DEBUG: target offers:", len(offer_to_target_sku), "target skus:", len(target_skus))
    print("SVD DEBUG: clusters for details:", len(used_clusters))

    def _num(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    def _orders_total(metrics):
        oc = (metrics or {}).get("orders_count", 0)
        if isinstance(oc, dict):
            return _num(oc.get("total", 0))
        return _num(oc)

    offer_weighted_time = {}
    offer_orders = {}
    offer_simple_sum = {}
    offer_simple_cnt = {}
    details_url = "https://api-seller.ozon.ru/v1/analytics/average-delivery-time/details"

    for period in svd_periods:
        need_offers = {o for o in target_offers if o not in offer_simple_cnt}
        if not need_offers:
            break
        print(f"SVD DEBUG: period={period} need_offers={len(need_offers)}")
        for cid in used_clusters:
            try:
                offset = 0
                loaded = 0
                while True:
                    body = {
                        "cluster_id": int(cid),
                        "filters": {"delivery_schema": "ALL", "supply_period": period},
                        "limit": 1000,
                        "offset": offset
                    }
                    resp = post_with_retry(details_url, HEADERS, body, timeout=60)
                    resp.raise_for_status()
                    rows = resp.json().get("data", [])
                    if not rows:
                        break

                    for row in rows:
                        item = row.get("item") or {}
                        metrics = row.get("metrics") or {}
                        t = _num(metrics.get("average_delivery_time"))
                        if t <= 0:
                            continue

                        offer_id = str(item.get("offer_id") or "").strip()
                        sku_raw = item.get("sku")
                        offer_by_sku = None
                        if sku_raw not in (None, ""):
                            try:
                                offer_by_sku = sku_to_offer.get(int(sku_raw))
                            except Exception:
                                pass

                        matched_offer = None
                        if offer_id in need_offers:
                            matched_offer = offer_id
                        elif offer_by_sku in need_offers:
                            matched_offer = offer_by_sku

                        if not matched_offer:
                            continue

                        orders = _orders_total(metrics)
                        if orders > 0:
                            offer_weighted_time[matched_offer] = offer_weighted_time.get(matched_offer, 0.0) + (t * orders)
                            offer_orders[matched_offer] = offer_orders.get(matched_offer, 0.0) + orders
                        offer_simple_sum[matched_offer] = offer_simple_sum.get(matched_offer, 0.0) + t
                        offer_simple_cnt[matched_offer] = offer_simple_cnt.get(matched_offer, 0) + 1
                        loaded += 1

                    if len(rows) < 1000:
                        break
                    offset += 1000
                print(f"SVD DEBUG: period={period} cluster={cid} loaded rows={loaded}")
            except Exception as e:
                print(f"SVD DEBUG: period={period} cluster={cid} failed ->", e)

    found = len(offer_simple_cnt)
    print("SVD DEBUG: matrix stats", f"offers_found={found}", f"offers_total={len(target_offers)}")

    offer_final = {}
    for offer in offer_simple_cnt:
        if offer_orders.get(offer, 0) > 0:
            offer_final[offer] = offer_weighted_time[offer] / offer_orders[offer]
        else:
            offer_final[offer] = offer_simple_sum[offer] / offer_simple_cnt[offer]

    if offer_final:
        store_avg = int(math.ceil((sum(offer_final.values()) / len(offer_final)) - 1e-9))
    else:
        store_avg = 0

    result = {}
    fallback_offers = 0
    for offer in target_offers:
        if offer in offer_final:
            svd_val = int(math.ceil(offer_final[offer] - 1e-9))
            result[offer] = svd_val
            print(f"SVD DEBUG: offer={offer} svd={svd_val} source=details_all_clusters orders={int(offer_orders.get(offer, 0))}")
            continue
        fallback_offers += 1
        result[offer] = store_avg
        print(f"SVD DEBUG: offer={offer} svd={store_avg} source=store_fallback_offer_missing")

    print("SVD DEBUG: fallback offers:", fallback_offers)
    return result

# ----------------------------------------------------
# ГЛАВНАЯ ЛОГИКА
# ----------------------------------------------------

def get_current_hour_bounds():
    # текущее время берём по локальному (НСК), затем переводим в UTC
    now_local = datetime.now(NSK)
    now_utc = now_local.astimezone(timezone.utc).replace(minute=0, second=0, microsecond=0)

    # предыдущий полностью завершённый час UTC
    start_utc = now_utc - timedelta(hours=1)
    # Закрываем интервал на последней секунде часа (без .999, только .000)
    end_utc = now_utc - timedelta(seconds=1)

    # для таблицы отображаем МСК
    start_msk = start_utc.astimezone(MSK)
    end_msk = (end_utc + timedelta(seconds=1)).astimezone(MSK)
    date_label = start_msk.strftime("%d.%m.%Y")
    hour_label = f"{start_msk.strftime('%H:00')}-{end_msk.strftime('%H:00')}"

    return {
        "start_utc": start_utc,
        "end_utc": end_utc,
        "start_msk": start_msk,
        "end_msk": end_msk,
        "date_label": date_label,
        "hour_label": hour_label,
    }


def get_current_hour_msk_utc_range():
    bounds = get_current_hour_bounds()
    return (
        to_ozon_utc_iso(bounds["start_utc"]),
        to_ozon_utc_iso(bounds["end_utc"]),
        bounds["date_label"],
        bounds["hour_label"],
    )

def main():

    sheet = connect_sheet()
    structure = read_table_structure(sheet)
    manual_metrics = load_manual_metrics(offer_to_sku)
    hourly_economics = load_hourly_economics_from_calc(offer_to_sku)

    offers = list({s[0] for s in structure})

    bounds = get_current_hour_bounds()
    date_from = to_ozon_utc_iso(bounds["start_utc"])
    date_to = to_ozon_utc_iso(bounds["end_utc"])
    date_label = bounds["date_label"]
    hour_label = bounds["hour_label"]
    state_date = bounds["start_msk"].strftime("%Y-%m-%d")

    day_start_msk = bounds["start_msk"].replace(hour=0, minute=0, second=0, microsecond=0)
    day_start_utc_iso = to_ozon_utc_iso(day_start_msk)
    current_cumulative_to_iso = to_ozon_utc_iso(bounds["end_utc"])

    postings = get_postings(date_from, date_to)
    current_cumulative_totals = get_finance_totals(state_date)
    prev_state = load_hourly_state()
    global _LAST_POSTINGS_CACHE
    _LAST_POSTINGS_CACHE = postings

    print(f"POSTINGS COUNT: {len(postings)}")
    if postings:
        print(f"SAMPLE KEYS: {list(postings[0].keys())}")
    print(f"CURRENT CUMULATIVE TOTALS KEYS: {sorted(current_cumulative_totals.keys())}")
    print(f"STATE DEBUG: prev_state={prev_state}")

    sales_stats = extract_hour_sales_stats(postings)
    prices = {
        offer: item["avg_price"]
        for offer, item in sales_stats.items()
        if item.get("qty", 0) > 0
    }
    hour_summary = build_hour_summary_from_state(
        current_cumulative_totals,
        prev_state,
        state_date,
        hour_label,
        bounds["end_msk"] + timedelta(seconds=1),
    )
    print(f"PRICES FOUND: {len(prices)}")
    ovh = get_dimensions(offers)
    # если за час нет продаж — ничего не пишем
    if not postings:
        print(f"EMPTY HOUR {date_label} {hour_label} -> skip write")
        return

    # колонку создаём/находим только когда есть данные
    col = get_or_create_date_column(sheet, date_label, hour_label)

    svd = get_svd_per_offer(offers)

    values = []
    loss_alerts = []
    low_margin_alerts = []
    alert_margin_found = 0
    alert_margin_missing_offers = set()
    found_cost_share = 0
    missing_cost_share_offers = set()
    missing_hourly_economics_offers = set()
    sales_offers_count = len(sales_stats)
    hourly_margin_computed = 0
    hourly_margin_missing_data = 0

    for offer, metric in structure:
        if metric == "Цена (по акции)":
            v = prices.get(offer, "")
            values.append(v if v != 0 else 0)

        elif metric == "Доля затрат":
            manual_data = manual_metrics.get(offer)
            if manual_data and manual_data.get("cost_share") != "":
                values.append(manual_data["cost_share"])
                found_cost_share += 1
            else:
                values.append("")
                if offer not in missing_cost_share_offers:
                    print(f"MANUAL SHEET WARNING: no cost_share for offer {offer}")
                    missing_cost_share_offers.add(offer)

        elif metric == "ОВХ":
            v = ovh.get(offer, "")
            if isinstance(v, (int, float)) and v > 0:
                values.append(math.ceil(v))
            else:
                values.append(v if v != 0 else 0)

        elif metric == "СВД":
            v = svd.get(offer)
            values.append(v if v is not None else 0)

        elif metric == "%":
            sale = sales_stats.get(offer)
            economics = hourly_economics.get(offer)
            margin_percent_value = ""

            if not sale or sale.get("qty", 0) <= 0 or sale.get("revenue", 0) <= 0:
                values.append("")
            else:
                qty_hour = sale.get("qty", 0)
                revenue_hour = sale.get("revenue", 0.0)
                unit_margin_rub = economics.get("unit_margin_rub") if economics else None

                if unit_margin_rub is None:
                    values.append("")
                    hourly_margin_missing_data += 1
                    if offer not in missing_hourly_economics_offers:
                        print(
                            f"HOURLY MARGIN WARNING: no unit_margin_rub for offer {offer}, "
                            f"qty={qty_hour}, revenue={revenue_hour}"
                        )
                        missing_hourly_economics_offers.add(offer)
                else:
                    margin_rub_hour = qty_hour * unit_margin_rub
                    margin_percent_value = (margin_rub_hour / revenue_hour) * 100 if revenue_hour > 0 else ""
                    values.append(margin_percent_value)
                    hourly_margin_computed += 1

            # --- ALERT CHECK ---
            if margin_percent_value != "":
                try:
                    margin_percent_number = float(margin_percent_value)
                    alert_margin_found += 1
                    margin_ratio = margin_percent_number / 100
                    margin_pct = round(margin_percent_number, 1)

                    if margin_ratio < 0:
                        loss_alerts.append(f"{offer} — {margin_pct}%")
                    elif margin_ratio < 0.10:
                        low_margin_alerts.append(f"{offer} — {margin_pct}%")
                except Exception:
                    print(f"ALERT WARNING: invalid hourly margin_percent for offer {offer}: {margin_percent_value}")
            else:
                if sale and sale.get("qty", 0) > 0 and offer not in alert_margin_missing_offers:
                    print(f"ALERT WARNING: no hourly margin_percent for offer {offer}")
                    alert_margin_missing_offers.add(offer)

        else:
            values.append("")

    total_margin_rub_hour = 0.0
    total_revenue_hour = 0.0
    margin_summary_missing_offers = []
    for offer, sale in sales_stats.items():
        qty_hour = sale.get("qty", 0) or 0
        revenue_hour = sale.get("revenue", 0.0) or 0.0
        if qty_hour <= 0 or revenue_hour <= 0:
            continue

        total_revenue_hour += revenue_hour
        economics = hourly_economics.get(offer) or {}
        unit_margin_rub = economics.get("unit_margin_rub")
        if unit_margin_rub is None:
            if len(margin_summary_missing_offers) < 10:
                margin_summary_missing_offers.append(offer)
            continue

        total_margin_rub_hour += qty_hour * unit_margin_rub

    if margin_summary_missing_offers:
        print(f"HOURLY SUMMARY WARNING: summary margin skipped offers without unit_margin_rub={margin_summary_missing_offers}")

    hour_summary["orders"] = total_revenue_hour
    if total_revenue_hour > 0 and hour_summary.get("commission", 0) <= 0:
        print("HOURLY SUMMARY WARNING: commission delta is empty while hourly sales exist -> use n/a in message")
        hour_summary["commission"] = None

    hour_summary["margin_rub"] = total_margin_rub_hour
    hour_summary["margin_pct"] = (total_margin_rub_hour / total_revenue_hour * 100) if total_revenue_hour > 0 else 0.0
    hour_summary["cost_share_pct"] = (
        (hour_summary["commission"] / total_revenue_hour) * 100
        if total_revenue_hour > 0 and hour_summary.get("commission") is not None
        else None
    )

    write_values(sheet, values, col)
    print(
        "HOURLY SUMMARY:",
        f"cost_share_found={found_cost_share}",
        f"calc_offers_loaded={len(hourly_economics)}",
        f"sales_offers_hour={sales_offers_count}",
        f"hourly_margin_computed={hourly_margin_computed}",
        f"hourly_margin_missing_data={hourly_margin_missing_data}",
        f"alert_margin_found={alert_margin_found}",
        f"hour_orders={safe_round(hour_summary.get('orders', 0), 2)}",
        f"hour_margin_rub={safe_round(hour_summary.get('margin_rub', 0), 2)}",
        f"hour_margin_pct={safe_round(hour_summary.get('margin_pct', 0), 2)}",
        f"hour_commission={safe_round(hour_summary.get('commission', 0), 2)}",
        f"hour_cost_share_pct={safe_round(hour_summary.get('cost_share_pct', 0), 2)}",
    )

    # ---- SEND HOURLY SUMMARY + ALERTS TO VK TEAMS ----
    msg = build_hourly_vk_message(hour_label, hour_summary, loss_alerts, low_margin_alerts)
    print("VK MESSAGE PREVIEW:\n" + msg)
    send_vk_teams_alert(msg)
    save_hourly_state(hour_summary.get("state", {}))


# ----------------------------------------------------

if __name__ == "__main__":
    main()
