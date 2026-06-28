import math
import os
import random
import re
import time
from datetime import datetime, timedelta, timezone
from io import BytesIO

import gspread
import requests
from dotenv import load_dotenv
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook

from constants import offer_to_sku


load_dotenv()

MSK = timezone(timedelta(hours=3))
DAYS_SHIFT = int(os.getenv("MARGIN_CLASSES_DAYS_SHIFT", "1"))
TARGET_DATE_OVERRIDE = os.getenv("TARGET_DATE", "").strip()
TARGET_DATE = TARGET_DATE_OVERRIDE or (
    datetime.now(MSK) - timedelta(days=DAYS_SHIFT)
).strftime("%Y-%m-%d")

GOOGLE_CREDS_PATH = os.getenv("GOOGLE_CREDS")

SOURCE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID_HOURLY_CALC")



SOURCE_WORKSHEET = "Калькулятор прибыли"

OUTPUT_SHEET_ID = os.getenv("SHEET_ID", "")

OUTPUT_WORKSHEET = "Лист1"

OZON_CLIENT_ID = os.getenv("OZON_CLIENT_ID")
OZON_API_KEY = os.getenv("OZON_API_KEY")

VK_TOKEN = os.getenv("VKTEAMS_BOT_TOKEN")
VK_API = (os.getenv("VKTEAMS_API") or "").rstrip("/")
VK_CHAT_ID = os.getenv("VK_CHAT_ID", "")
#VK_CHAT_ID = "user@example.com"
SKIP_SEND = os.getenv("MARGIN_CLASSES_SKIP_SEND", "").strip().lower() in {
    "1",
    "true",
    "yes",
    "y",
}

SUPPORTED_CLASSES = ("СМС", "ЖМС", "РФ")
# Подклассы из колонки «Тип» калькулятора, которые относятся к ЖМС
# (категорию ЖМС переразбили на детальные ярлыки). Ключи — в верхнем регистре
# без пробелов, как их отдаёт normalize_class.
CLASS_ALIASES = {
    "ГДС": "ЖМС",
    "КОНД": "ЖМС",
    "ГДМП": "ЖМС",
    "МЫЛО": "ЖМС",
    "ПОЛ": "ЖМС",
}
DEBUG_TOP_OFFERS = int(os.getenv("MARGIN_CLASSES_DEBUG_TOP_OFFERS", "10"))
CLASS_ROW_MAP = {
    "СМС": {"title": 2, "orders": 3, "margin": 4, "margin_pct": 5, "qty": 6},
    "ЖМС": {"title": 7, "orders": 8, "margin": 9, "margin_pct": 10, "qty": 11},
    "РФ": {"title": 12, "orders": 13, "margin": 14, "margin_pct": 15, "qty": 16},
}
METRIC_TITLES = {
    "orders": "Заказы",
    "margin": "Маржа",
    "margin_pct": "Маржинальность",
    "qty": "Кол-во",
}
RETRYABLE_GOOGLE_CODES = {408, 409, 429, 500, 502, 503, 504}


def require_env():
    missing = []
    for key, value in {
        "OZON_CLIENT_ID": OZON_CLIENT_ID,
        "OZON_API_KEY": OZON_API_KEY,
        "SOURCE_SHEET_ID or SOURCE_SHEET_NAME": SOURCE_SHEET_ID,
        "GOOGLE_CREDS": GOOGLE_CREDS_PATH,
    }.items():
        if not value:
            missing.append(key)
    if missing:
        raise RuntimeError("Missing required environment variables: " + ", ".join(missing))


def get_google_creds():
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    return ServiceAccountCredentials.from_json_keyfile_name(GOOGLE_CREDS_PATH, scope)


def get_gspread_client():
    return gspread.authorize(get_google_creds())


def _google_error_code(exc):
    response = getattr(exc, "response", None)
    if response is not None:
        try:
            return int(response.status_code)
        except Exception:
            pass
    text = str(exc)
    for code in RETRYABLE_GOOGLE_CODES:
        if f"'code': {code}" in text or f'"code": {code}' in text:
            return code
    return None


def google_call(label, func, *args, retries=6, base_wait=3, **kwargs):
    for attempt in range(1, retries + 1):
        try:
            return func(*args, **kwargs)
        except Exception as exc:
            code = _google_error_code(exc)
            retryable = code in RETRYABLE_GOOGLE_CODES or code is None
            if attempt == retries or not retryable:
                raise
            wait = min(base_wait * (2 ** (attempt - 1)), 60) + random.uniform(0, 2)
            print(f"GOOGLE RETRY {label}: {exc}. Sleep {wait:.1f}s ({attempt}/{retries})")
            time.sleep(wait)


def safe_request(url, body, timeout=60, max_retries=4, base_wait=5):
    headers = {
        "Client-Id": OZON_CLIENT_ID,
        "Api-Key": OZON_API_KEY,
        "Content-Type": "application/json",
    }
    for attempt in range(max_retries):
        try:
            resp = requests.post(url, headers=headers, json=body, timeout=timeout)
            if resp.status_code == 429 or 500 <= resp.status_code < 600:
                wait = base_wait * (2 ** attempt) + random.uniform(0, 2)
                print(f"OZON RETRY {url}: status={resp.status_code}. Sleep {wait:.1f}s")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp
        except requests.RequestException as exc:
            wait = base_wait * (2 ** attempt) + random.uniform(0, 2)
            print(f"OZON RETRY {url}: {exc}. Sleep {wait:.1f}s")
            time.sleep(wait)
    raise RuntimeError(f"Failed Ozon request after {max_retries} attempts: {url}")


def is_unsupported_document_error(exc):
    error_data = getattr(exc, "args", [{}])[0]
    if isinstance(error_data, dict):
        return "This operation is not supported for this document" in str(
            error_data.get("message", "")
        )
    return False


def download_drive_file_content(file_id):
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
    client = get_gspread_client()
    creds = get_google_creds()
    token = creds.get_access_token().access_token

    last_error = None
    for attempt in range(1, 7):
        try:
            resp = client.http_client.request("get", url)
            resp.raise_for_status()
            return resp.content
        except Exception as exc:
            last_error = exc
            wait = min(3 * (2 ** (attempt - 1)), 60) + random.uniform(0, 2)
            print(
                f"SOURCE WARNING: gspread download failed -> {exc}. "
                f"Sleep {wait:.1f}s ({attempt}/6)"
            )
            time.sleep(wait)

    for attempt in range(1, 7):
        try:
            resp = requests.get(
                url,
                headers={"Authorization": f"Bearer {token}"},
                timeout=60,
            )
            resp.raise_for_status()
            return resp.content
        except Exception as exc:
            last_error = exc
            wait = min(3 * (2 ** (attempt - 1)), 60) + random.uniform(0, 2)
            print(
                f"SOURCE WARNING: requests download failed -> {exc}. "
                f"Sleep {wait:.1f}s ({attempt}/6)"
            )
            time.sleep(wait)

    raise RuntimeError(f"SOURCE ERROR: failed to download xlsx from Drive -> {last_error}")


def read_xlsx_rows_by_id(file_id, worksheet_name):
    workbook = load_workbook(
        BytesIO(download_drive_file_content(file_id)),
        read_only=True,
        data_only=True,
    )
    if worksheet_name not in workbook.sheetnames:
        raise RuntimeError(f"SOURCE ERROR: worksheet {worksheet_name!r} not found in xlsx")
    ws = workbook[worksheet_name]
    return [list(row) for row in ws.iter_rows(min_col=1, max_col=200, values_only=True)]


def read_source_rows():
    client = get_gspread_client()
    try:
        if SOURCE_SHEET_ID:
            spreadsheet = google_call("open source by id", client.open_by_key, SOURCE_SHEET_ID)
        worksheet = google_call("open source worksheet", spreadsheet.worksheet, SOURCE_WORKSHEET)
        rows = google_call("read source values", worksheet.get_all_values)
        print("SOURCE DEBUG: source=gspread")
        return rows
    except gspread.exceptions.APIError as exc:
        if SOURCE_SHEET_ID and is_unsupported_document_error(exc):
            print("SOURCE DEBUG: source=xlsx via Drive API")
            return read_xlsx_rows_by_id(SOURCE_SHEET_ID, SOURCE_WORKSHEET)
        raise


def normalize_header_name(value):
    if value in (None, ""):
        return ""
    text = str(value).replace("\n", " ").replace("\xa0", " ").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def parse_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip()
    if not raw:
        return None
    raw = (
        raw.replace("\xa0", "")
        .replace(" ", "")
        .replace("%", "")
        .replace("₽", "")
        .replace("руб.", "")
        .replace("руб", "")
        .replace("шт.", "")
        .replace("шт", "")
        .replace(",", ".")
    )
    try:
        return float(raw)
    except Exception:
        return None


def normalize_offer(raw_value):
    if raw_value in (None, ""):
        return ""
    if isinstance(raw_value, (int, float)):
        value = str(int(raw_value)) if float(raw_value).is_integer() else str(raw_value)
    else:
        value = str(raw_value).strip()
    value = value.replace("\xa0", " ").strip()
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"(?i)\s*ИР\s*$", "", value).strip()
    value = re.sub(r"(?<=\d)\.0+$", "", value)
    match = re.match(r"^\s*(\d+)(?:/\d+)?(?:\s*([A-Za-zА-Яа-яЁё]+))?\s*$", value)
    if not match:
        return ""
    suffix = match.group(2)
    return f"{match.group(1)} {suffix.upper()}".strip() if suffix else match.group(1)


def build_offer_lookup_candidates(raw_value):
    if raw_value in (None, ""):
        return []
    value = str(raw_value).replace("\xa0", " ").strip().upper()
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"(?i)\s*ИР\s*$", "", value).strip()
    candidates = []

    def add(candidate):
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    add(normalize_offer(value))
    add(normalize_offer(re.sub(r"(?<=\d)/\d+", "", value).strip()))
    base = re.match(r"^\s*(\d+)", value)
    if base:
        add(base.group(1))
    return candidates


def resolve_offer_key(raw_value):
    for candidate in build_offer_lookup_candidates(raw_value):
        if candidate in offer_to_sku:
            return candidate
    return ""


def normalize_class(value):
    text = str(value or "").replace("\xa0", " ").strip().upper()
    text = re.sub(r"\s+", "", text)
    text = CLASS_ALIASES.get(text, text)
    return text if text in SUPPORTED_CLASSES else ""


def find_source_columns(rows):
    aliases = {
        "class": {"тип"},
        "article": {"артикул"},
        "margin_rub": {"маржа fbo, руб"},
    }
    for row_idx, row in enumerate(rows):
        normalized = [normalize_header_name(cell) for cell in row]
        found = {}
        for key, values in aliases.items():
            for col_idx, cell in enumerate(normalized):
                if cell in values:
                    found[key] = col_idx
                    break
        if all(key in found for key in aliases):
            print(
                "SOURCE DEBUG:",
                f"header_row={row_idx + 1}",
                f"class_col={found['class'] + 1}",
                f"article_col={found['article'] + 1}",
                f"margin_rub_col={found['margin_rub'] + 1}",
            )
            return row_idx, found
    raise RuntimeError("SOURCE ERROR: headers 'Тип', 'Артикул', 'Маржа FBO, руб' not found")


def load_article_classes_and_margins():
    rows = read_source_rows()
    header_idx, columns = find_source_columns(rows)

    data = {}
    ignored_classes = {}
    duplicate_offers = 0
    missing_margin = []

    for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        raw_article = row[columns["article"]] if len(row) > columns["article"] else ""
        offer = resolve_offer_key(raw_article)
        if not raw_article or not offer:
            continue

        cls = normalize_class(row[columns["class"]] if len(row) > columns["class"] else "")
        if not cls:
            raw_cls = str(row[columns["class"]] if len(row) > columns["class"] else "").strip()
            ignored_classes[raw_cls or "EMPTY"] = ignored_classes.get(raw_cls or "EMPTY", 0) + 1
            print(f"SOURCE IGNORE: row={row_idx} offer={offer} unsupported_class={raw_cls!r}")
            continue

        margin_rub = parse_number(row[columns["margin_rub"]] if len(row) > columns["margin_rub"] else "")
        if margin_rub is None:
            missing_margin.append(offer)
            print(f"SOURCE WARNING: row={row_idx} offer={offer} class={cls} missing margin_rub")
            continue

        if offer in data:
            duplicate_offers += 1
            print(f"SOURCE WARNING: duplicate offer {offer} at row {row_idx}, overwrite")

        data[offer] = {"class": cls, "unit_margin_rub": margin_rub}

    print(
        "SOURCE SUMMARY:",
        f"loaded_offers={len(data)}",
        f"duplicates={duplicate_offers}",
        f"ignored_classes={ignored_classes}",
        f"missing_margin={len(missing_margin)}",
    )
    return data


def utc_day_range(date_str):
    dt_msk = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=MSK)
    utc_from = (dt_msk - timedelta(hours=3)).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    utc_to = (
        dt_msk + timedelta(days=1) - timedelta(hours=3) - timedelta(seconds=1)
    ).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    return utc_from, utc_to


def build_sku_to_offer_index():
    result = {}
    for offer, sku in offer_to_sku.items():
        result[str(sku)] = str(offer)
        try:
            result[int(sku)] = str(offer)
        except Exception:
            pass
    return result


def get_postings(date_from, date_to):
    url = "https://api-seller.ozon.ru/v3/posting/fbo/list"
    postings = []
    cursor = ""
    while True:
        body = {
            "sort_dir": "ASC",
            "filter": {"since": date_from, "to": date_to},
            "limit": 100,
            "cursor": cursor,
            "translit": True,
            "with": {
                "analytics_data": True,
                "financial_data": True,
                "legal_info": False,
            },
        }
        resp = safe_request(url, body)
        r = resp.json()
        batch = r.get("postings", [])
        if not batch:
            break
        postings.extend(batch)
        cursor = r.get("cursor", "")
        if not r.get("has_next", False):
            break
    return postings


def extract_sales_stats(postings):
    stats = {}
    sku_to_offer = build_sku_to_offer_index()

    def map_offer(sku):
        if sku in sku_to_offer:
            return sku_to_offer[sku]
        try:
            return sku_to_offer.get(int(sku))
        except Exception:
            return sku_to_offer.get(str(sku))

    def num(value):
        if isinstance(value, dict):
            value = value.get("amount", value.get("value", 0))
        try:
            return float(value)
        except Exception:
            return 0.0

    def qty(value):
        try:
            parsed = int(value)
            return parsed if parsed > 0 else 0
        except Exception:
            return 0

    def add_sale(offer, price, amount):
        if not offer or price <= 0 or amount <= 0:
            return
        item = stats.setdefault(offer, {"qty": 0, "revenue": 0.0})
        item["qty"] += amount
        item["revenue"] += price * amount

    for posting in postings:
        used_financial = set()
        for product in posting.get("financial_data", {}).get("products", []) or []:
            offer = map_offer(product.get("product_id"))
            price = num(product.get("price"))
            amount = qty(product.get("quantity"))
            if offer and price > 0 and amount > 0:
                add_sale(offer, price, amount)
                used_financial.add(offer)

        for product in posting.get("products", []) or []:
            offer = map_offer(product.get("sku"))
            if not offer or offer in used_financial:
                continue
            add_sale(offer, num(product.get("price")), qty(product.get("quantity")))

    return stats


def build_class_totals(sales_stats, article_data):
    totals = {
        cls: {"orders": 0.0, "margin": 0.0, "margin_pct": 0.0, "qty": 0}
        for cls in SUPPORTED_CLASSES
    }
    control = {
        "source_orders": 0.0,
        "source_qty": 0,
        "accounted_orders": 0.0,
        "accounted_qty": 0,
        "ignored_orders": 0.0,
        "ignored_qty": 0,
    }
    ignored = {
        "missing_article_data": [],
        "unsupported_class": [],
    }
    class_offer_details = {cls: [] for cls in SUPPORTED_CLASSES}

    for offer, sale in sales_stats.items():
        qty_value = sale.get("qty", 0) or 0
        revenue = sale.get("revenue", 0.0) or 0.0
        if qty_value <= 0 or revenue <= 0:
            continue

        control["source_orders"] += revenue
        control["source_qty"] += qty_value

        data = article_data.get(offer)
        if not data:
            ignored["missing_article_data"].append(offer)
            control["ignored_orders"] += revenue
            control["ignored_qty"] += qty_value
            print(f"CLASS IGNORE: offer={offer} reason=missing_class_or_margin")
            continue

        cls = data.get("class")
        if cls not in totals:
            ignored["unsupported_class"].append(offer)
            control["ignored_orders"] += revenue
            control["ignored_qty"] += qty_value
            print(f"CLASS IGNORE: offer={offer} reason=unsupported_class class={cls!r}")
            continue

        totals[cls]["orders"] += revenue
        margin_value = qty_value * data["unit_margin_rub"]
        totals[cls]["margin"] += margin_value
        totals[cls]["qty"] += qty_value
        control["accounted_orders"] += revenue
        control["accounted_qty"] += qty_value
        class_offer_details[cls].append(
            {
                "offer": offer,
                "qty": qty_value,
                "revenue": revenue,
                "unit_margin_rub": data["unit_margin_rub"],
                "margin": margin_value,
                "margin_pct": margin_value / revenue * 100 if revenue > 0 else 0.0,
            }
        )

    for cls, item in totals.items():
        item["margin_pct"] = (
            item["margin"] / item["orders"] * 100 if item["orders"] > 0 else 0.0
        )

    print(
        "CLASS SUMMARY:",
        f"sales_offers={len(sales_stats)}",
        f"source_orders={round(control['source_orders'], 2)}",
        f"source_qty={control['source_qty']}",
        f"accounted_orders={round(control['accounted_orders'], 2)}",
        f"accounted_qty={control['accounted_qty']}",
        f"ignored_orders={round(control['ignored_orders'], 2)}",
        f"ignored_qty={control['ignored_qty']}",
        f"missing_article_data={len(ignored['missing_article_data'])}",
        f"unsupported_class={len(ignored['unsupported_class'])}",
    )
    for cls in SUPPORTED_CLASSES:
        item = totals[cls]
        avg_unit_margin = item["margin"] / item["qty"] if item["qty"] > 0 else 0.0
        print(
            "MARGIN CLASS DEBUG:",
            f"class={cls}",
            f"orders={round(item['orders'], 2)}",
            f"qty={item['qty']}",
            f"margin={round(item['margin'], 2)}",
            f"margin_pct={round(item['margin_pct'], 2)}",
            f"avg_unit_margin={round(avg_unit_margin, 2)}",
        )
        details = sorted(
            class_offer_details[cls],
            key=lambda row: abs(row["margin"]),
            reverse=True,
        )[:DEBUG_TOP_OFFERS]
        for row in details:
            print(
                "MARGIN OFFER DEBUG:",
                f"class={cls}",
                f"offer={row['offer']}",
                f"qty={row['qty']}",
                f"revenue={round(row['revenue'], 2)}",
                f"unit_margin={round(row['unit_margin_rub'], 2)}",
                f"margin={round(row['margin'], 2)}",
                f"margin_pct={round(row['margin_pct'], 2)}",
            )
    return totals, ignored, control


def connect_output_sheet():
    client = get_gspread_client()
    if OUTPUT_SHEET_ID:
        spreadsheet = google_call("open output by id", client.open_by_key, OUTPUT_SHEET_ID)
    try:
        return google_call("open output worksheet", spreadsheet.worksheet, OUTPUT_WORKSHEET)
    except gspread.exceptions.WorksheetNotFound:
        print(f"OUTPUT WARNING: worksheet {OUTPUT_WORKSHEET!r} not found, use first worksheet")
        return spreadsheet.sheet1


def parse_sheet_date(value):
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass
    return None


def get_or_create_date_column(sheet, date_str):
    header = google_call("read output row 1", sheet.row_values, 1)
    target_date = datetime.strptime(date_str, "%Y-%m-%d").date()

    for idx, value in enumerate(header, start=1):
        if parse_sheet_date(value) == target_date:
            print(f"OUTPUT DEBUG: date {date_str} already exists in col={idx}")
            return idx

    insert_pos = None
    for idx, value in enumerate(header, start=1):
        existing = parse_sheet_date(value)
        if existing and existing > target_date:
            insert_pos = idx
            break
    if insert_pos is None:
        insert_pos = max(len(header), 1) + 1

    google_call("insert output col", sheet.insert_cols, [[]], col=insert_pos)
    print(f"OUTPUT DEBUG: inserted date {date_str} at col={insert_pos}")
    return insert_pos


def ensure_output_labels(sheet):
    labels = [[""] for _ in range(16)]
    labels[0][0] = "Дата"
    for cls, rows in CLASS_ROW_MAP.items():
        labels[rows["title"] - 1][0] = cls
        for metric, title in METRIC_TITLES.items():
            labels[rows[metric] - 1][0] = title
    google_call(
        "write output labels",
        sheet.update,
        values=labels,
        range_name="A1:A16",
    )


def write_totals(sheet, date_str, totals):
    ensure_output_labels(sheet)
    col = get_or_create_date_column(sheet, date_str)
    values = [[""] for _ in range(16)]
    display_date = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d.%m.%Y")
    values[0][0] = display_date

    for cls, rows in CLASS_ROW_MAP.items():
        item = totals[cls]
        values[rows["orders"] - 1][0] = round(item["orders"], 2)
        values[rows["margin"] - 1][0] = round(item["margin"], 2)
        values[rows["margin_pct"] - 1][0] = round(item["margin_pct"], 2)
        values[rows["qty"] - 1][0] = int(item["qty"])

    start = gspread.utils.rowcol_to_a1(1, col)
    end = gspread.utils.rowcol_to_a1(16, col)
    google_call(
        "write output totals",
        sheet.update,
        values=values,
        range_name=f"{start}:{end}",
        value_input_option="USER_ENTERED",
    )
    return col


def read_totals_by_date(sheet, date_str):
    header = google_call("read output row 1", sheet.row_values, 1)
    target = datetime.strptime(date_str, "%Y-%m-%d").date()
    col = None
    for idx, value in enumerate(header, start=1):
        if parse_sheet_date(value) == target:
            col = idx
            break
    if not col:
        return None

    raw_values = google_call("read output col", sheet.col_values, col)
    result = {}
    for cls, rows in CLASS_ROW_MAP.items():
        result[cls] = {}
        for metric in ("orders", "margin", "margin_pct", "qty"):
            idx = rows[metric] - 1
            value = raw_values[idx] if idx < len(raw_values) else ""
            result[cls][metric] = parse_number(value) or 0.0
    return result


def format_money(value):
    return f"{float(value):,.0f} ₽".replace(",", " ")


def format_qty(value):
    return f"{int(round(float(value))):,}".replace(",", " ")


def format_percent(value):
    return f"{float(value):.1f} %"


def format_delta(value):
    if value is None:
        return "нет данных"
    if value > 0:
        return f"+{abs(value):.1f}%"
    if value < 0:
        return f"-{abs(value):.1f}%"
    return "0.0%"


def calc_dynamics(now, past):
    if past is None or past == 0:
        return None
    return (now - past) / past * 100


def calc_pp_delta(now, past):
    if past is None:
        return None
    return now - past


def metric_delta_text(metric, now_value, past_totals, cls):
    if not past_totals:
        return "нет данных"
    past_value = past_totals.get(cls, {}).get(metric)
    if metric == "margin_pct":
        delta = calc_pp_delta(now_value, past_value)
        return f"{delta:+.1f} п.п." if delta is not None else "нет данных"
    return format_delta(calc_dynamics(now_value, past_value))


def build_message(date_str, totals, sheet, ignored, control):
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    prev_day = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
    prev_week = (date_obj - timedelta(days=7)).strftime("%Y-%m-%d")
    prev_month = (date_obj - timedelta(days=30)).strftime("%Y-%m-%d")

    day_totals = read_totals_by_date(sheet, prev_day)
    week_totals = read_totals_by_date(sheet, prev_week)
    month_totals = read_totals_by_date(sheet, prev_month)

    display_date = date_obj.strftime("%d.%m.%Y")
    lines = [f"📊Маржа по классам Ozon (заказы) — {display_date}"]

    for cls in SUPPORTED_CLASSES:
        item = totals[cls]
        lines.extend(
            [
                "",
                cls,
                (
                    f"Заказы: {format_money(item['orders'])} "
                    f"(д/д: {metric_delta_text('orders', item['orders'], day_totals, cls)} · "
                    f"д/нед: {metric_delta_text('orders', item['orders'], week_totals, cls)} · "
                    f"д/мес: {metric_delta_text('orders', item['orders'], month_totals, cls)})"
                ),
                (
                    f"Маржа: {format_money(item['margin'])} "
                    f"(д/д: {metric_delta_text('margin', item['margin'], day_totals, cls)} · "
                    f"д/нед: {metric_delta_text('margin', item['margin'], week_totals, cls)} · "
                    f"д/мес: {metric_delta_text('margin', item['margin'], month_totals, cls)})"
                ),
                (
                    f"Маржинальность: {format_percent(item['margin_pct'])} "
                    f"(д/д: {metric_delta_text('margin_pct', item['margin_pct'], day_totals, cls)} · "
                    f"д/нед: {metric_delta_text('margin_pct', item['margin_pct'], week_totals, cls)} · "
                    f"д/мес: {metric_delta_text('margin_pct', item['margin_pct'], month_totals, cls)})"
                ),
                (
                    f"Кол-во: {format_qty(item['qty'])} шт "
                    f"(д/д: {metric_delta_text('qty', item['qty'], day_totals, cls)} · "
                    f"д/нед: {metric_delta_text('qty', item['qty'], week_totals, cls)} · "
                    f"д/мес: {metric_delta_text('qty', item['qty'], month_totals, cls)})"
                ),
            ]
        )

    missing_count = len(ignored.get("missing_article_data", []))
    unsupported_count = len(ignored.get("unsupported_class", []))
    if missing_count or unsupported_count or control.get("ignored_qty", 0):
        lines.extend(
            [
                "",
                (
                    f"Контроль: всего заказано {format_money(control.get('source_orders', 0))} "
                    f"/ {format_qty(control.get('source_qty', 0))} шт; "
                    f"учтено {format_money(control.get('accounted_orders', 0))} "
                    f"/ {format_qty(control.get('accounted_qty', 0))} шт"
                ),
                (
                    f"Не учтено: {format_money(control.get('ignored_orders', 0))} "
                    f"/ {format_qty(control.get('ignored_qty', 0))} шт; "
                    f"без класса/маржи — {missing_count}, неподдержанный класс — {unsupported_count}"
                ),
            ]
        )
    return "\n".join(lines)


def send_to_vk(text):
    if SKIP_SEND:
        print("VK SKIP: MARGIN_CLASSES_SKIP_SEND is enabled")
        return
    if not VK_API or not VK_TOKEN or not VK_CHAT_ID:
        print("VK WARNING: env vars are not fully configured, skip send")
        return
    url = f"{VK_API}/messages/sendText"
    resp = requests.post(
        url,
        data={"token": VK_TOKEN, "chatId": VK_CHAT_ID, "text": text},
        timeout=20,
    )
    print(f"VK STATUS: {resp.status_code}")
    print(f"VK RESPONSE: {resp.text}")
    resp.raise_for_status()


def main():
    require_env()
    print(f"TARGET DATE: {TARGET_DATE}")

    article_data = load_article_classes_and_margins()
    date_from, date_to = utc_day_range(TARGET_DATE)
    print(f"OZON WINDOW: {date_from} -> {date_to}")

    postings = get_postings(date_from, date_to)
    print(f"POSTINGS COUNT: {len(postings)}")

    sales_stats = extract_sales_stats(postings)
    totals, ignored, control = build_class_totals(sales_stats, article_data)

    sheet = connect_output_sheet()
    col = write_totals(sheet, TARGET_DATE, totals)
    print(f"OUTPUT UPDATED: date={TARGET_DATE} col={col}")

    message = build_message(TARGET_DATE, totals, sheet, ignored, control)
    print("VK MESSAGE PREVIEW:\n" + message)
    send_to_vk(message)


if __name__ == "__main__":
    main()
